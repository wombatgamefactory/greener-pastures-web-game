# -*- coding: utf-8 -*-
"""Extract packages/data/data/cards.json from the designer's spreadsheet.

Run:  python tools/extract_cards.py <path-to-sheet.xlsm>
  or:  GP_SHEET=<path> python tools/extract_cards.py

The sheet is DELIBERATELY not committed and its path is DELIBERATELY not
hard-coded here: this repo is published, the sheet is private design material,
and its filename names the game under its pitch title. Supply the path.

The sheet (worksheet "cards") is the upstream source of truth for card data;
packages/data/data/cards.json is what the game reads. Re-run this when a new
sheet version lands, then diff the JSON.

Authoritative columns are the @cost1..@cost6 icon columns, NOT the human-readable
"Build Cost" text column - the latter is lossy (it prints a total, "2 resources",
and says nothing about which crops satisfy it).

SCHEMA 2 (design changes v31, 02/09/2026). Every card is a single FLAT object -
there is no `faces` key and no `handSize`, `upgradeCostCoins` or `coins` anywhere.

  * A suit has THREE starters - Barn (ref 1), Farmstead (ref 2), Notice Board
    (ref 3) - plus an 18-card shuffled deck (refs 4..21). 105 cards in total.
  * Starters are SINGLE-FACED. v31 deleted all fifteen upgraded "U" rows from the
    sheet along with the starter-upgrade rule itself, so a `U` row reaching this
    script means an old sheet was passed in, and it is reported as an error
    rather than absorbed. Starters are not bought: they print no build cost and
    carry no cost icons, and their printed VP is 0.
  * The Barn prints NO text at all (it is simply where cards ready for delivery
    are stored), so its `abilityText` is the empty string.
  * The Farmstead prints one end-game line, "Game end: 1 VP for each <CROP> card
    you have built."; it holds no cards, so it prints no threshold.
  * The Notice Board is the only loadable starter: threshold 2, wild activation,
    and its VISITOR line grants that suit's door action.
  * COINS ARE GONE from the game. A coin cost icon or a `£` in any card text is a
    stale sheet, and both are reported as errors.

Keys are camelCase here so the JSON is the TypeScript shape with no mapping layer.
"""
import collections
import hashlib
import json
import os
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "packages" / "data" / "data" / "cards.json"

# Columns are resolved by HEADER TEXT, never pinned by letter, and a missing
# header is a hard error naming it. tools/make_web_assets.py already does this and
# says why; here is what it costs not to. A `Notes` column at R was deleted at
# sheet v22 and everything to its right shifted one place left, which repointed
# the cost block at @cost2..@cost6 PLUS `total_cost`. The crash that followed was
# luck: `icons = [i for i in icons if i]` drops falsy values, so a `total_cost` of
# 0 would have been filtered away silently and every card's build cost written
# one icon short, with no error at all. Only a non-zero total reached `icon_cost`
# and raised on an int.
COL_HEADERS = {
    "cardback": "Cardback", "suit": "Suit", "ref": "Ref", "type": "Type",
    "name": "Name", "buildCostText": "Build Cost", "activation": "Activation Cost",
    "threshold": "Threshold", "vp": "VP", "effect": "Activation Effect",
    "ability": "Ability",
}
COST_HEADERS = ["@cost%d" % i for i in range(1, 7)]
TOTAL_COST_HEADER = "total_cost"
CARD_NUM_HEADER = "Card#"

TYPES = {"Starter": "starter", "Tier 1": "tier1", "Tier 2": "tier2",
         "Tier 3": "tier3", "Power": "power", "Endgame": "endgame"}

# The three starter slots, keyed by printed ref number (1/2/3).
STARTER_SLOT_BY_NUM = {1: "barn", 2: "farmstead", 3: "noticeboard"}

# --- Pinned mechanical stats ---------------------------------------------
# The Notice Board prints its threshold and activation type as columns, so both
# are READ from the sheet; the pin below exists so a silent sheet edit is caught
# rather than absorbed. v31 set it to 2 (it was 5 through the whole v14 era, and
# the sheet had already drifted to 2 ahead of the rule - the ledger flagged that
# gap, and this pin is what closes it).
NOTICE_BOARD_THRESHOLD = 2

# Every Power and Endgame card costs 2 cards of its OWN suit from v31; no coins,
# no wilds. Pinned for the same reason as the Notice Board threshold.
POWER_ENDGAME_COST = {"suit": 2, "wild": 0}

# Terms that named systems the game no longer has. Any of them in a card's text
# means the sheet is behind the rules, which is worth an error rather than a
# shrug: the sheet is the single source of truth for wording, so a stale string
# here is a stale string on the printed card.
RETIRED_TERMS = ("Hired Hand", "Hired Worker", "Working Week", "County Show",
                 "Hand size", "buy at market", "£")

# NO TEXT OVERRIDE LIVES HERE, deliberately. The sheet is the single source of
# truth for card wording, so the web game and the physical game cannot drift apart
# on rules text. The reference implementation carried an override on the Orchard
# Farmstead on the belief that the sheet was stale; that turned out to be a
# different resolution of the same design question rather than staleness, so the
# override is gone and the sheet wins. A ruling that changes wording is applied to
# the sheet first, then re-extracted.

# A warning containing any of these is a hard failure (non-zero exit), not a note.
# Each marks a SILENT sheet edit - a value the game depends on that changed
# upstream without anyone saying so - rather than a cosmetic quibble.
FATAL_MARKERS = ("expected", "cannot be loaded", "Notice Board", "unrecognised starter",
                 "coin cost icon", "upgraded 'U' row", "retired term", "starter")


def resolve_columns(ws):
    """Map every column this script reads onto its letter, by matching header text.

    Returns (col, cost_cols, total_cost, card_num). Exits naming the header if the
    sheet does not carry it, because a `None` here is indistinguishable from an
    empty cell and would be absorbed into the extract rather than reported.
    """
    header = {}
    for i in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=i).value
        if v is not None and str(v).strip():
            header.setdefault(str(v).strip(), get_column_letter(i))

    missing = [h for h in list(COL_HEADERS.values()) + COST_HEADERS
               + [TOTAL_COST_HEADER, CARD_NUM_HEADER] if h not in header]
    if missing:
        sys.exit("sheet worksheet 'cards' is missing %d required column header(s): %s"
                 % (len(missing), ", ".join(missing)))

    return ({k: header[h] for k, h in COL_HEADERS.items()},
            [header[h] for h in COST_HEADERS],
            header[TOTAL_COST_HEADER], header[CARD_NUM_HEADER])


def sheet_path():
    """The sheet lives outside this repo. Path from argv[1] or $GP_SHEET."""
    positional = [x for x in sys.argv[1:] if not x.startswith("-")]
    raw = positional[0] if positional else os.environ.get("GP_SHEET")
    if not raw:
        sys.exit("usage: python tools/extract_cards.py <path-to-sheet.xlsm>\n"
                 "   or: set GP_SHEET to the sheet's path")
    p = Path(raw).expanduser()
    if not p.is_file():
        sys.exit("no such sheet: %s" % p)
    return p


def out_path():
    """Where the JSON lands. Defaults to the game's own cards.json.

    `--out` exists for the card sheet renderer, which extracts a WORKING COPY
    from the shared Google Sheet on every render. That copy must not overwrite
    `cards.json`: the sheet is allowed to be mid-thought, and a half-drawn
    experiment must not become the game's baseline behind anybody's back. The
    `--out` branch at the bottom of main() is what lets the renderer draw such a
    sheet anyway - refusing would make it useless exactly when a designer most
    wants to see the change.
    """
    if "--out" in sys.argv:
        return Path(sys.argv[sys.argv.index("--out") + 1]).expanduser()
    return OUT


# The printed TRIGGER PREFIX, stripped. From v17 the sheet prefixes a card's
# effect line with `GROW: ` or `ACTION: ` (Orchard and Vegetable carry it; the
# earlier suits do not, and Wheat's rows are unprefixed to this day). It is a
# LAYOUT label naming the trigger, not part of the effect sentence, and the
# trigger is already carried structurally in `abilityTrigger` - so keeping it
# would put the same fact in two places and make one suit's text read
# differently from another's for no reason. NOT a text override: nothing is
# rewritten, a prefix the art prints in its own right is simply not duplicated
# into the effect string.
TRIGGER_PREFIX = re.compile(r"^\s*(?:GROW|ACTION)\s*:\s*", re.IGNORECASE)


def clean(v):
    """Cells encode art line-breaks as a literal backslash-n."""
    if v is None:
        return None
    text = re.sub(r"[ \t]+\n", "\n", str(v).replace("\\n", "\n").strip())
    text = TRIGGER_PREFIX.sub("", text)
    return text or None


def icon_cost(icons):
    """suit_wheat.png -> 'suit'; suit_wild.png -> 'wild'.

    `coin_front.png` is counted separately so the caller can REPORT it. Coins left
    the game at v31, so a coin bar on the sheet is a stale row, and the count is
    never written into the JSON - `buildCost` carries `suit` and `wild` only.
    """
    c = collections.Counter()
    for i in icons:
        if i == "coin_front.png":
            c["coins"] += 1
        elif i == "suit_wild.png":
            c["wild"] += 1
        elif i.startswith("suit_"):
            c["suit"] += 1
        else:
            raise ValueError("unrecognised cost icon: %s" % i)
    return c


TRIGGER_PATTERNS = [
    ("autoHarvest", r"automatically harvests"),
    ("onHarvest", r"when (?:this card is )?harvested"),
    # The Deliver trigger vocabulary has two levels, selected by printed phrase.
    # "When you Deliver..." fires on island claims AND balloon moves;
    # "When you Deliver to the island..." fires on island claims only.
    ("onDeliverIsland", r"when you deliver to the island"),
    ("onDeliver", r"when you deliver(?! to the island)"),
]
# The two surcharge triggers that used to live here, `harvestSurcharge` and
# `activationSurcharge`, both matched "must pay £1 to ...". Coins are gone, no
# card carries that wording any more, and a pattern that can never match is a
# pattern nobody maintains - so they went with the currency.


def triggers_for(card_type, text, threshold=None, activation=None):
    """Detected trigger keywords. Compound/absent detections are flagged, not guessed.

    Keyword detection over the printed text, NOT a resolved ruling. needsDesignReview
    just means 0 or >1 triggers matched and a human must look.

    The one STRUCTURAL detection is `action`, and it is structural because it has
    to be: an ACTION card prints no prefix on the sheet, but it prints no threshold
    and no activation type either, so it can be neither grown nor sown. A tier card
    with text and no way to be activated fires as a main action instead.

    Starters return no trigger. That is unchanged from schema 1, and it is worth
    naming because the v31 Farmstead is an end-game scorer: its VP is counted by
    the engine's scoring pass off the card's identity, not off this array.
    """
    if card_type == "power":
        low = (text or "").lower()
        deliver = [name for name, pat in TRIGGER_PATTERNS
                   if name.startswith("onDeliver") and re.search(pat, low)]
        return ["passive"] + deliver, False
    if card_type == "endgame":
        return ["gameEnd"], False
    if card_type == "starter" or not text:
        return [], False
    if threshold is None and activation is None:
        return ["action"], False

    low = text.lower()
    found = [name for name, pat in TRIGGER_PATTERNS if re.search(pat, low)]
    starts_on_harvest = re.match(
        r"^\s*(this building automatically|when (?:this card is )?harvested)", low)
    if not starts_on_harvest:
        found.insert(0, "onActivate")
    ambiguous = len(found) != 1
    return found, ambiguous


def check_starter(card, warnings):
    """Check the printed stats of one starter against what the rules require.

    The Notice Board is the only loadable starter (threshold 2 / wild); the Barn
    and the Farmstead hold no cards, so they must print neither threshold nor
    activation, and the Barn must print no text at all.
    """
    slot, cid = card["slot"], card["id"]

    if slot == "noticeboard":
        if (card["threshold"], card["activationType"]) != (NOTICE_BOARD_THRESHOLD, "wild"):
            warnings.append(
                "%s: Notice Board must be threshold %d / wild activation "
                "(sheet says %r / %r)"
                % (cid, NOTICE_BOARD_THRESHOLD, card["threshold"], card["activationType"]))
        if not card["abilityText"]:
            warnings.append("%s: Notice Board prints no VISITOR line" % cid)
        return

    if card["threshold"] is not None or card["activationType"] is not None:
        warnings.append("%s: the %s cannot be loaded, so it must print no threshold and "
                        "no activation type (sheet says %r / %r)"
                        % (cid, slot, card["threshold"], card["activationType"]))

    if slot == "barn" and card["abilityText"]:
        warnings.append("%s: the Barn prints no text from v31, but the sheet still says %r"
                        % (cid, card["abilityText"]))
    if slot == "farmstead" and not card["abilityText"]:
        warnings.append("%s: expected the Farmstead's end-game line, found no text" % cid)


def g_cardnum(ws, card_num_col, r):
    """The `Card#` cell as a string. Its own helper only because the U-row guard
    reads it before the row is otherwise parsed."""
    return str(ws["%s%d" % (card_num_col, r)].value or "").strip()


def main():
    xlsm = sheet_path()
    digest = hashlib.sha256(xlsm.read_bytes()).hexdigest()
    wb = openpyxl.load_workbook(xlsm, data_only=True)
    ws = wb["cards"]
    col, cost_cols, total_cost_col, card_num_col = resolve_columns(ws)

    cards, warnings, uncached = {}, [], []

    for r in range(2, ws.max_row + 1):
        def g(c, _r=r):
            return ws["%s%d" % (col[c], _r)].value

        suit, ref, ctype, name = g("suit"), g("ref"), g("type"), g("name")
        if not (suit and ref and ctype and name):
            if ctype == "Free Port":
                warnings.append("r%d: skipped '%s' row -- the Aerodrome is a per-player "
                                "board, not a deck card; it lives in data/aerodrome.json"
                                % (r, ctype))
            continue
        if ctype not in TYPES:
            warnings.append("r%d: skipped unknown type %r" % (r, ctype))
            continue

        # v31 deleted the starter-upgrade rule and, with it, the fifteen `NNU`
        # rows that carried the upgraded faces. Schema 1 paired the two rows on
        # `Ref` and emitted a `faces` object; there is nothing to pair now, so a
        # surviving `U` row can only mean a pre-v31 sheet was passed in. Reported
        # rather than absorbed: silently dropping it would extract a sheet the
        # caller thinks is current, and silently keeping it would put a phantom
        # 16th card in a suit.
        ref = str(ref).strip()
        if ref.upper().endswith("U") or str(g_cardnum(ws, card_num_col, r)).upper().endswith("U"):
            warnings.append("r%d %s: upgraded 'U' row found -- starters are single-faced "
                            "from v31 and all fifteen U rows were deleted from the sheet. "
                            "This looks like a pre-v31 sheet." % (r, ref))
            continue

        ctype = TYPES[ctype]
        suit = suit.lower()
        icons = [ws["%s%d" % (c, r)].value for c in cost_cols]
        icons = [i for i in icons if i]
        # `total_cost` is the sheet's own =COUNTA() over the icon block, read here
        # as a cross-check that the icon columns still line up (the v22 column
        # shift that this file's header comment describes was caught by exactly
        # this). It is a CACHED formula value, and openpyxl drops the cache when
        # it saves - so a sheet whose last writer was a script, not Excel, has an
        # empty cache on every row. That is not a data error and must not shout:
        # counted once, reported once, and the cross-check simply cannot run.
        total = ws["%s%d" % (total_cost_col, r)].value
        if total is None or total == "":
            uncached.append(ref)
        elif total != len(icons):
            warnings.append("r%d %s: totalCost=%s but %d cost icons" % (r, ref, total, len(icons)))
        cost = icon_cost(icons)
        if cost["coins"]:
            warnings.append("r%d %s: %d coin cost icon(s) -- coins were removed from the "
                            "game at v31 and never reach the JSON" % (r, ref, cost["coins"]))

        vp = g("vp") or 0
        threshold = g("threshold")
        activation = clean(g("activation"))

        # RULED (Dean, 2026-08-13): the two text columns mean something. A card
        # with an ACTIVATION power prints in `Activation Effect`; a card with a
        # STATIC effect prints in `Ability`. So the Barn and the Farmstead, both
        # passive, live in `Ability`, while the Notice Board - the one loadable
        # starter - keeps its VISITOR text in `Activation Effect`. Read whichever
        # is filled, for every card type alike.
        text = clean(g("effect")) or clean(g("ability"))

        stale = [t for t in RETIRED_TERMS if t.lower() in (text or "").lower()]
        if stale:
            warnings.append("r%d %s: retired term(s) %s in the printed text %r"
                            % (r, ref, ", ".join(repr(s) for s in stale), text))

        if ctype == "starter":
            slot = STARTER_SLOT_BY_NUM.get(int(re.sub(r"\D", "", ref) or 0))
            if not slot:
                warnings.append("r%d %s: unrecognised starter ref" % (r, ref))
            if icons or clean(g("buildCostText")):
                warnings.append("r%d %s: a starter is not bought, so it must print no build "
                                "cost and carry no cost icons (found %s / %r)"
                                % (r, ref, dict(cost), clean(g("buildCostText"))))
            if vp:
                warnings.append("r%d %s: starter printed VP is %s, expected 0" % (r, ref, vp))
            cards[(suit, ref)] = {
                "id": ref, "suit": suit, "type": "starter", "slot": slot,
                "name": clean(name), "inDeck": False, "enabled": True,
                "buildCost": None,
                "activationType": activation, "threshold": threshold,
                "printedVp": 0,
                "abilityText": text or "",
                "abilityTrigger": [], "needsDesignReview": False,
            }
            check_starter(cards[(suit, ref)], warnings)
            continue

        trigger, ambiguous = triggers_for(ctype, text, threshold, activation)
        build_cost = {"suit": cost["suit"], "wild": cost["wild"]}
        if ctype in ("power", "endgame") and build_cost != POWER_ENDGAME_COST:
            warnings.append("r%d %s: %s card build cost %s, expected %s (2 cards of its "
                            "own suit, ruled 02/09/2026)"
                            % (r, ref, ctype, build_cost, POWER_ENDGAME_COST))
        cards[(suit, ref)] = {
            "id": ref, "suit": suit, "type": ctype, "name": clean(name), "inDeck": True,
            "enabled": True,
            "buildCost": build_cost,
            "activationType": activation, "threshold": threshold, "printedVp": vp,
            "abilityText": text or "",
            "abilityTrigger": trigger,
            "needsDesignReview": ambiguous,
        }

    ordered = sorted(cards.values(), key=lambda c: (c["suit"], int(re.sub(r"\D", "", c["id"]))))

    # --- validation -------------------------------------------------------
    by_suit = collections.Counter(c["suit"] for c in ordered)
    by_type = collections.Counter((c["suit"], c["type"]) for c in ordered)
    expect = {"starter": 3, "tier1": 5, "tier2": 4, "tier3": 3, "power": 3, "endgame": 3}
    for suit in by_suit:
        for t, n in expect.items():
            if by_type[(suit, t)] != n:
                warnings.append("%s: expected %d %s, found %d" % (suit, n, t, by_type[(suit, t)]))
        deck = sum(1 for c in ordered if c["suit"] == suit and c["inDeck"])
        if deck != 18:
            warnings.append("%s: expected 18 shuffled deck cards, found %d" % (suit, deck))
        slots = sorted(c["slot"] for c in ordered if c["suit"] == suit and not c["inDeck"])
        if slots != ["barn", "farmstead", "noticeboard"]:
            warnings.append("%s: expected the three starter slots, found %s" % (suit, slots))
    if len(ordered) != 105:
        warnings.append("expected 105 cards total (5 x (3 starters + 18 deck)), found %d"
                        % len(ordered))
    if uncached:
        note = ("%d row(s) carry no cached total_cost, so the icon-block cross-check "
                "could not run on them. Expected when the sheet was last written by a "
                "script: openpyxl drops formula caches on save, and Excel refills them "
                "the next time the file is opened and saved." % len(uncached))
        if len(uncached) == len(ordered):
            print("note: " + note)
        else:
            warnings.append(note)

    doc = {
        "meta": {
            "schemaVersion": 2,
            "kind": "generated",
            "generatedBy": "tools/extract_cards.py",
            "sourceSheet": "worksheet 'cards' of the designer spreadsheet (kept outside this repo)",
            "sourceSha256": digest,
            "notes": [
                "buildCost is derived from the @cost1..@cost6 icon columns, which are"
                " authoritative; the sheet's 'Build Cost' text column is lossy (it prints"
                " a total and does not say which crops satisfy it).",
                "A suit has THREE starters - Barn (ref 1), Farmstead (ref 2), Notice Board"
                " (ref 3) - plus the 18-card deck (refs 4..21). 105 cards total"
                " (15 starters + 90 deck).",
                "Schema 2 (design changes v31, 02/09/2026) is FLAT: every card is one"
                " object with no nested per-side block. Starters lost their upgraded side"
                " along with the starter-upgrade rule; the hand-limit field went with the"
                " hand limit, the starter-upgrade price went with the upgrade, and the"
                " money term of a build cost went with money itself. A build cost is now"
                " exactly {suit, wild}. Anything the schema-1 shape carried and this one"
                " does not is deliberate, not missing.",
                "Starters are not bought: buildCost is null and printedVp is 0 on all"
                " fifteen. The Barn prints no text (abilityText is the empty string), the"
                " Farmstead prints one end-game line scoring its own crop, and the Notice"
                " Board is the only loadable starter (threshold 2, wild activation), whose"
                " VISITOR line grants that suit's door action.",
                "Every Power and Endgame card costs 2 cards of its own suit"
                " ({suit: 2, wild: 0}). Money does not exist in the game from v31, so no"
                " monetary term reaches this file; a money icon still printed on the"
                " sheet's cost bar is reported as an error rather than absorbed.",
                "No card text is ever rewritten here. The sheet is the single source of"
                " truth for wording, so the web game and the physical game cannot drift"
                " apart on rules text; the tuning overlay carries numbers and flags only."
                " A ruling that changes wording is applied to the sheet first, then"
                " re-extracted.",
                "Every card carries `enabled: true`. It is a tuning-overlay flag, not a"
                " printed property: switching a card off is how a paired comparison run"
                " asks whether the game is better without it.",
                "abilityTrigger is keyword detection over the printed text, not a resolved"
                " ruling. needsDesignReview=true means 0 or >1 triggers matched, and those"
                " cards are re-read by hand when their handlers are written. Starters carry"
                " no trigger at all - the Farmstead's end-game VP is counted by the"
                " engine's scoring pass off the card's identity, not off this array.",
                "The one STRUCTURAL trigger is `action`: a tier card with printed text but"
                " no threshold and no activation type can be neither grown nor sown, so its"
                " text fires as a MAIN ACTION instead - taken in place of Draw, Build, Grow,"
                " Harvest or Deliver. The sheet prints no prefix for it, which is why it is"
                " read off the shape of the card rather than off a keyword.",
                "The Deliver trigger has TWO keys: onDeliver ('When you Deliver...', fires on"
                " island claims and balloon moves) and onDeliverIsland ('When you Deliver to"
                " the island...', island claims only).",
            ],
        },
        "suits": sorted(by_suit),
        "catalogue": ordered,
    }
    out = out_path()
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(doc, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    review = [c["id"] for c in ordered if c["needsDesignReview"]]
    shuffled = sum(c["inDeck"] for c in ordered)
    print("wrote %s: %d cards (%d shuffled, %d starters)"
          % (out, len(ordered), shuffled, len(ordered) - shuffled))
    print("source sha256: %s" % digest)
    print("per suit: %s" % dict(by_suit))
    print("ambiguous abilityTrigger (%d): %s" % (len(review), " ".join(review)))
    if warnings:
        print("\n%d warning(s):" % len(warnings))
        for w in warnings:
            print("  - %s" % w)
    fatal = [w for w in warnings if any(m in w for m in FATAL_MARKERS)]
    if fatal and out != OUT:
        # FATAL_MARKERS guard the GAME's invariants, and they must keep doing
        # that for `cards.json`. A `--out` copy is a proof render of whatever
        # the sheet currently says, and the sheet is allowed to be mid-thought:
        # a Notice Board at 3 is an experiment being drawn, not corrupt data.
        # Refusing to draw it would make the renderer useless exactly when a
        # designer most wants to see the change.
        print("")
        print("%d of those break a rule the GAME enforces. Writing anyway: %s is a "
              "proof copy, not the baseline." % (len(fatal), out.name))
        return 0
    return 1 if fatal else 0


if __name__ == "__main__":
    sys.exit(main())
